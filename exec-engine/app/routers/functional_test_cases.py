"""
功能测试用例库：标准模板 CRUD，以及 Excel / XMind 导入。
"""
from __future__ import annotations

import io
import json
import re
import zipfile
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session

from ..database import get_db
from ..models.api_mgmt import FunctionalTestCase
from ..schemas.api_mgmt import (
    FunctionalTestCaseCreate,
    FunctionalTestCaseListPayload,
    FunctionalTestCaseListResponse,
    FunctionalTestCaseResponse,
    FunctionalTestCaseUpdate,
    ImportResultPayload,
    ResponseModel,
)

router = APIRouter(prefix="/functional-test-cases", tags=["Functional Test Cases"])

# --- 表头别名（中英文）→ 内部字段 ---
_EXCEL_ALIASES: Dict[str, str] = {
    "module": "module",
    "模块": "module",
    "所属模块": "module",
    "用例模块": "module",
    "case_code": "case_code",
    "编号": "case_code",
    "用例编号": "case_code",
    "case id": "case_code",
    "title": "title",
    "用例标题": "title",
    "标题": "title",
    "用例名称": "title",
    "name": "title",
    "priority": "priority",
    "优先级": "priority",
    "category": "category",
    "用例类型": "category",
    "类型": "category",
    "preconditions": "preconditions",
    "前置条件": "preconditions",
    "前置": "preconditions",
    "steps": "steps",
    "测试步骤": "steps",
    "步骤": "steps",
    "操作步骤": "steps",
    "expected_result": "expected_result",
    "预期结果": "expected_result",
    "期望结果": "expected_result",
    "remark": "remark",
    "备注": "remark",
    "说明": "remark",
    "status": "status",
    "状态": "status",
}


def _norm_header(cell: Any) -> Optional[str]:
    if cell is None:
        return None
    s = str(cell).strip().lower()
    if not s:
        return None
    return _EXCEL_ALIASES.get(s) or _EXCEL_ALIASES.get(str(cell).strip())


def _parse_steps_text(text: str) -> List[Dict[str, Any]]:
    if not text or not str(text).strip():
        return []
    lines = [ln.strip() for ln in str(text).splitlines() if ln.strip()]
    out: List[Dict[str, Any]] = []
    for i, line in enumerate(lines, 1):
        action = re.sub(r"^\d+[\.\、\)]\s*", "", line)
        action = re.sub(r"^[（(]?\d+[）)]\s*", "", action)
        out.append({"order": i, "action": action, "expected": ""})
    return out


def _attached(topic: Dict[str, Any]) -> List[Dict[str, Any]]:
    ch = topic.get("children") or {}
    if isinstance(ch, dict):
        return list(ch.get("attached") or [])
    return []


def _leaves_from_topic(topic: Dict[str, Any], ancestors: List[str]) -> List[Dict[str, Any]]:
    title = (topic.get("title") or "").strip() or "未命名"
    path = ancestors + [title]
    kids = _attached(topic)
    if not kids:
        if len(path) >= 2:
            mod = " / ".join(path[:-1])
            return [
                {
                    "module": mod,
                    "title": path[-1],
                    "steps": None,
                    "expected_result": None,
                }
            ]
        return []
    rows: List[Dict[str, Any]] = []
    for k in kids:
        rows.extend(_leaves_from_topic(k, path))
    return rows


def _parse_xmind_bytes(raw: bytes) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        names = [n for n in zf.namelist() if n.endswith(".json") and "content" in n.replace("\\", "/").lower()]
        if not names:
            names = [n for n in zf.namelist() if n.endswith("content.json")]
        if not names:
            raise ValueError("XMind 压缩包中未找到 content.json")
        for name in names:
            try:
                data = json.loads(zf.read(name).decode("utf-8"))
            except Exception:
                continue
            sheets: List[Any]
            if isinstance(data, list):
                sheets = data
            else:
                sheets = [data]
            for sheet in sheets:
                if not isinstance(sheet, dict):
                    continue
                rt = sheet.get("rootTopic")
                if isinstance(rt, dict):
                    out.extend(_leaves_from_topic(rt, []))
    if not out:
        raise ValueError("未能从 XMind 中解析出任何叶子主题，请确认脑图结构")
    return out


def _row_to_model_fields(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    title = (row.get("title") or "").strip()
    if not title:
        return None
    steps_raw = row.get("steps")
    steps: Optional[List[Dict[str, Any]]] = None
    if isinstance(steps_raw, str) and steps_raw.strip():
        steps = _parse_steps_text(steps_raw)
    elif isinstance(steps_raw, list):
        steps = steps_raw

    pr = (row.get("priority") or "P2").strip().upper()
    if pr not in ("P0", "P1", "P2", "P3"):
        pr = "P2"

    st = (row.get("status") or "draft").strip().lower()
    if st not in ("draft", "ready", "deprecated"):
        st = "draft"

    return {
        "module": (row.get("module") or "").strip()[:512],
        "case_code": (row.get("case_code") or "").strip()[:128] or None,
        "title": title[:512],
        "priority": pr,
        "category": (row.get("category") or "functional").strip()[:64] or "functional",
        "preconditions": (row.get("preconditions") or "").strip() or None,
        "steps": steps,
        "expected_result": (row.get("expected_result") or "").strip() or None,
        "remark": (row.get("remark") or "").strip() or None,
        "status": st,
    }


def _read_excel_rows(content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise HTTPException(status_code=500, detail="服务器未安装 openpyxl，无法导入 Excel") from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Excel 为空"]

    col_index: Dict[int, str] = {}
    for i, cell in enumerate(header_row):
        key = _norm_header(cell)
        if key:
            col_index[i] = key

    if "title" not in col_index.values():
        return [], ["Excel 首行需包含「用例标题」或 title 列"]

    data_rows: List[Dict[str, Any]] = []
    for line in rows_iter:
        item: Dict[str, Any] = {}
        for idx, key in col_index.items():
            if idx < len(line):
                val = line[idx]
                if val is not None and not isinstance(val, str):
                    val = str(val).strip()
                item[key] = val
        if any(v not in (None, "") for v in item.values()):
            data_rows.append(item)
    return data_rows, []


@router.get("", response_model=FunctionalTestCaseListResponse)
def list_functional_cases(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    module: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(FunctionalTestCase)
    if module:
        q = q.filter(FunctionalTestCase.module.contains(module.strip()))
    if keyword:
        kw = f"%{keyword.strip()}%"
        q = q.filter(
            (FunctionalTestCase.title.like(kw))
            | (FunctionalTestCase.case_code.like(kw))
            | (FunctionalTestCase.remark.like(kw))
        )
    total = q.count()
    items = (
        q.order_by(FunctionalTestCase.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return FunctionalTestCaseListResponse(
        data=FunctionalTestCaseListPayload(items=items, total=total)
    )


@router.post("", response_model=FunctionalTestCaseResponse)
def create_functional_case(body: FunctionalTestCaseCreate, db: Session = Depends(get_db)):
    row = FunctionalTestCase(**body.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return FunctionalTestCaseResponse(data=row)


@router.get("/{case_id}", response_model=FunctionalTestCaseResponse)
def get_functional_case(case_id: int, db: Session = Depends(get_db)):
    row = db.query(FunctionalTestCase).filter(FunctionalTestCase.id == case_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="用例不存在")
    return FunctionalTestCaseResponse(data=row)


@router.patch("/{case_id}", response_model=FunctionalTestCaseResponse)
def update_functional_case(
    case_id: int, body: FunctionalTestCaseUpdate, db: Session = Depends(get_db)
):
    row = db.query(FunctionalTestCase).filter(FunctionalTestCase.id == case_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="用例不存在")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(row, k, v)
    db.commit()
    db.refresh(row)
    return FunctionalTestCaseResponse(data=row)


@router.delete("/{case_id}", response_model=ResponseModel)
def delete_functional_case(case_id: int, db: Session = Depends(get_db)):
    row = db.query(FunctionalTestCase).filter(FunctionalTestCase.id == case_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="用例不存在")
    db.delete(row)
    db.commit()
    return ResponseModel(msg="deleted")


@router.post("/import/excel", response_model=ResponseModel)
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx / .xlsm 文件")
    raw = await file.read()
    rows, errs = _read_excel_rows(raw)
    if errs and not rows:
        raise HTTPException(status_code=400, detail=errs[0])

    imported = 0
    skipped = 0
    errors: List[str] = []
    for i, r in enumerate(rows, start=2):
        fields = _row_to_model_fields(r)
        if not fields:
            skipped += 1
            continue
        try:
            db.add(FunctionalTestCase(**fields))
            imported += 1
        except Exception as e:
            errors.append(f"第{i}行: {e}")
    db.commit()
    return ResponseModel(data=ImportResultPayload(imported=imported, skipped=skipped, errors=errors[:20]))


@router.post("/import/xmind", response_model=ResponseModel)
async def import_xmind(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename or not file.filename.lower().endswith(".xmind"):
        raise HTTPException(status_code=400, detail="请上传 .xmind 文件")
    raw = await file.read()
    try:
        parsed = _parse_xmind_bytes(raw)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except zipfile.BadZipFile as e:
        raise HTTPException(status_code=400, detail="无效的 XMind 压缩包") from e

    imported = 0
    for item in parsed:
        fields = _row_to_model_fields(
            {
                "module": item.get("module") or "",
                "title": item.get("title") or "",
                "steps": item.get("steps"),
                "expected_result": item.get("expected_result"),
            }
        )
        if not fields:
            continue
        db.add(FunctionalTestCase(**fields))
        imported += 1
    db.commit()
    return ResponseModel(data=ImportResultPayload(imported=imported, skipped=0, errors=[]))
