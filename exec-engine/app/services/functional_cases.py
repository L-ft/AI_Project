from __future__ import annotations

import io
import json
import re
import zipfile
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException, UploadFile

from ..models.api_mgmt import FunctionalTestCase
from ..repositories.functional_cases import FunctionalTestCaseRepository


EXCEL_ALIASES: Dict[str, str] = {
    "module": "module",
    "模块": "module",
    "所属模块": "module",
    "用例模块": "module",
    "case_code": "case_code",
    "编号": "case_code",
    "用例编号": "case_code",
    "case id": "case_code",
    "title": "title",
    "用例标题": "title",
    "标题": "title",
    "用例名称": "title",
    "name": "title",
    "priority": "priority",
    "优先级": "priority",
    "category": "category",
    "用例类型": "category",
    "类型": "category",
    "preconditions": "preconditions",
    "前置条件": "preconditions",
    "前置": "preconditions",
    "steps": "steps",
    "测试步骤": "steps",
    "步骤": "steps",
    "操作步骤": "steps",
    "expected_result": "expected_result",
    "预期结果": "expected_result",
    "期望结果": "expected_result",
    "remark": "remark",
    "备注": "remark",
    "说明": "remark",
    "status": "status",
    "状态": "status",
}


def _normalize_header(cell: Any) -> Optional[str]:
    if cell is None:
        return None
    raw = str(cell).strip()
    if not raw:
        return None
    return EXCEL_ALIASES.get(raw.lower()) or EXCEL_ALIASES.get(raw)


def _parse_steps_text(text: str) -> List[Dict[str, Any]]:
    if not text or not str(text).strip():
        return []

    items: List[Dict[str, Any]] = []
    lines = [line.strip() for line in str(text).splitlines() if line.strip()]
    for index, line in enumerate(lines, start=1):
        action = re.sub(r"^\d+[\.\、\)]\s*", "", line)
        action = re.sub(r"^[（(]?\d+[)）]\s*", "", action)
        items.append({"order": index, "action": action, "expected": ""})
    return items


def _attached_topics(topic: Dict[str, Any]) -> List[Dict[str, Any]]:
    children = topic.get("children") or {}
    if isinstance(children, dict):
        return list(children.get("attached") or [])
    return []


def _leaf_rows_from_topic(topic: Dict[str, Any], ancestors: List[str]) -> List[Dict[str, Any]]:
    title = (topic.get("title") or "").strip() or "Untitled"
    path = ancestors + [title]
    children = _attached_topics(topic)
    if not children:
        if len(path) < 2:
            return []
        return [
            {
                "module": " / ".join(path[:-1]),
                "title": path[-1],
                "steps": None,
                "expected_result": None,
            }
        ]

    rows: List[Dict[str, Any]] = []
    for child in children:
        rows.extend(_leaf_rows_from_topic(child, path))
    return rows


def _parse_xmind_bytes(raw: bytes) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        names = [
            name
            for name in archive.namelist()
            if name.endswith(".json") and "content" in name.replace("\\", "/").lower()
        ]
        if not names:
            names = [name for name in archive.namelist() if name.endswith("content.json")]
        if not names:
            raise ValueError("content.json was not found in the XMind archive")

        for name in names:
            try:
                data = json.loads(archive.read(name).decode("utf-8"))
            except Exception:
                continue

            sheets = data if isinstance(data, list) else [data]
            for sheet in sheets:
                if not isinstance(sheet, dict):
                    continue
                root_topic = sheet.get("rootTopic")
                if isinstance(root_topic, dict):
                    rows.extend(_leaf_rows_from_topic(root_topic, []))

    if not rows:
        raise ValueError("No test case leaf nodes were parsed from the XMind file")
    return rows


def _row_to_model_fields(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    title = str(row.get("title") or "").strip()
    if not title:
        return None

    raw_steps = row.get("steps")
    steps: Optional[List[Dict[str, Any]]] = None
    if isinstance(raw_steps, str) and raw_steps.strip():
        steps = _parse_steps_text(raw_steps)
    elif isinstance(raw_steps, list):
        steps = raw_steps

    priority = str(row.get("priority") or "P2").strip().upper()
    if priority not in ("P0", "P1", "P2", "P3"):
        priority = "P2"

    status = str(row.get("status") or "draft").strip().lower()
    if status not in ("draft", "ready", "deprecated"):
        status = "draft"

    category = str(row.get("category") or "functional").strip() or "functional"

    return {
        "module": str(row.get("module") or "").strip()[:512],
        "case_code": (str(row.get("case_code") or "").strip()[:128] or None),
        "title": title[:512],
        "priority": priority,
        "category": category[:64],
        "preconditions": (str(row.get("preconditions") or "").strip() or None),
        "steps": steps,
        "expected_result": (str(row.get("expected_result") or "").strip() or None),
        "remark": (str(row.get("remark") or "").strip() or None),
        "status": status,
    }


def _read_excel_rows(content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required to import Excel files") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Excel file is empty"]

    column_map: Dict[int, str] = {}
    for index, cell in enumerate(header_row):
        key = _normalize_header(cell)
        if key:
            column_map[index] = key

    if "title" not in column_map.values():
        return [], ["Excel header must include a title column"]

    rows: List[Dict[str, Any]] = []
    for raw_row in rows_iter:
        item: Dict[str, Any] = {}
        for index, key in column_map.items():
            if index >= len(raw_row):
                continue
            value = raw_row[index]
            if value is not None and not isinstance(value, str):
                value = str(value).strip()
            item[key] = value
        if any(value not in (None, "") for value in item.values()):
            rows.append(item)
    return rows, []


class FunctionalTestCaseService:
    def __init__(self, repository: FunctionalTestCaseRepository):
        self.repository = repository

    def _serialize_case(self, row: FunctionalTestCase) -> Dict[str, Any]:
        return {
            "id": row.code,
            "code": row.code,
            "module": row.module,
            "case_code": row.case_code,
            "caseCode": row.case_code,
            "title": row.title,
            "priority": row.priority,
            "category": row.category,
            "preconditions": row.preconditions,
            "steps": row.steps,
            "expected_result": row.expected_result,
            "expectedResult": row.expected_result,
            "remark": row.remark,
            "status": row.status,
        }

    def _get_case_entity(self, case_ref: object) -> FunctionalTestCase:
        row = self.repository.get_by_ref(case_ref)
        if not row:
            raise HTTPException(status_code=404, detail="Functional test case not found")
        return row

    def list_cases(
        self,
        page: int,
        page_size: int,
        module: Optional[str] = None,
        keyword: Optional[str] = None,
    ) -> Tuple[List[Dict[str, Any]], int]:
        items, total = self.repository.list_paginated(page, page_size, module, keyword)
        return [self._serialize_case(item) for item in items], total

    def create_case(self, body) -> Dict[str, Any]:
        entity = FunctionalTestCase(**body.model_dump())
        return self._serialize_case(self.repository.save(entity))

    def get_case(self, case_id: object) -> Dict[str, Any]:
        return self._serialize_case(self._get_case_entity(case_id))

    def update_case(self, case_id: object, body) -> Dict[str, Any]:
        row = self._get_case_entity(case_id)
        for key, value in body.model_dump(exclude_unset=True).items():
            setattr(row, key, value)
        return self._serialize_case(self.repository.save(row))

    def delete_case(self, case_id: object) -> None:
        row = self._get_case_entity(case_id)
        self.repository.delete(row)

    async def import_excel(self, file: UploadFile) -> Dict[str, Any]:
        if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail="Please upload a .xlsx or .xlsm file")

        raw = await file.read()
        rows, errors = _read_excel_rows(raw)
        if errors and not rows:
            raise HTTPException(status_code=400, detail=errors[0])

        entities: List[FunctionalTestCase] = []
        skipped = 0
        for row in rows:
            fields = _row_to_model_fields(row)
            if not fields:
                skipped += 1
                continue
            entities.append(FunctionalTestCase(**fields))

        if entities:
            self.repository.add_all(entities)
            self.repository.commit()

        return {
            "imported": len(entities),
            "skipped": skipped,
            "errors": errors[:20],
        }

    async def import_xmind(self, file: UploadFile) -> Dict[str, Any]:
        if not file.filename or not file.filename.lower().endswith(".xmind"):
            raise HTTPException(status_code=400, detail="Please upload a .xmind file")

        raw = await file.read()
        try:
            parsed_rows = _parse_xmind_bytes(raw)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except zipfile.BadZipFile as exc:
            raise HTTPException(status_code=400, detail="Invalid XMind archive") from exc

        entities: List[FunctionalTestCase] = []
        for row in parsed_rows:
            fields = _row_to_model_fields(
                {
                    "module": row.get("module") or "",
                    "title": row.get("title") or "",
                    "steps": row.get("steps"),
                    "expected_result": row.get("expected_result"),
                }
            )
            if fields:
                entities.append(FunctionalTestCase(**fields))

        if entities:
            self.repository.add_all(entities)
            self.repository.commit()

        return {
            "imported": len(entities),
            "skipped": 0,
            "errors": [],
        }
